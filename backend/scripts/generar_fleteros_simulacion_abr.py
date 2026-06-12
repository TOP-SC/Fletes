"""
Genera Excel «Fletes solicitados sucursales» simulado para ABR 2026.

Usa casos Mundo 2 reales del Tango importado (abril) con el mismo formato
que el export Drive de mayo, para probar import + macheo de fleteros locales.

Uso:
  cd backend
  python scripts/generar_fleteros_simulacion_abr.py
  python scripts/generar_fleteros_simulacion_abr.py --limite 120 --salida ../data/mi_archivo.xlsx
"""

from __future__ import annotations

import argparse
import random
import re
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

_BACKEND = Path(__file__).resolve().parents[1]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from app.models import Envio
from app.services.fecha_utils import parse_fecha_tango
from app.services.mundo2_service import es_envio_mundo2
from app.services.remito_maestro import clave_agrupacion_caso

ROOT = _BACKEND.parent
DB_PATH = ROOT / "data" / "fletes.db"
DEFAULT_OUT = ROOT / "data" / "Fletes Solicitados sucursales ABR 2026 - Simulacion Logistica.xlsx"
MAYO_REF = Path(
    r"S:\Administración\TOP\LOG -  Envios Fletes 200326\Copia de Fletes Solicitados sucursales MAYO 2 - Logística.xlsx"
)

FLETEROS = [
    ("BLAS ANTONIO FERNÁNDEZ", 0.42),
    ("ARMANDO RIOS", 0.33),
    ("GAMA AGUSTIN JORGE EDUARDO", 0.15),
    ("OTROS", 0.10),
]

LOCALES = ["SF", "LA", "LO", "BE", "AV", "CA", "QU", "MO", "HA", "TU"]


def _elegir_fletero(rng: random.Random) -> str:
    x = rng.random()
    ac = 0.0
    for nombre, peso in FLETEROS:
        ac += peso
        if x <= ac:
            return nombre
    return FLETEROS[0][0]


def _fmt_fecha(d: date | None) -> str:
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


def _cliente_desde_envio(env: Envio) -> str:
    rs = (env.razon_social or "CLIENTE").strip()
    cod = ""
    raw = env.raw_json or ""
    m = re.search(r'"cod_cliente"\s*:\s*"([^"]+)"', raw, re.I)
    if m:
        cod = m.group(1).strip()
    if not cod and env.nro_pedido:
        cod = f"P{str(env.nro_pedido)[-5:]}"
    return f"{cod} - {rs}" if cod else rs


def _articulos_desde_grupo(lineas: list[Envio]) -> str:
    partes: list[str] = []
    vistos: set[str] = set()
    for e in lineas:
        ped = (e.nro_pedido or "").strip()
        if not ped or ped in vistos:
            continue
        vistos.add(ped)
        desc = (e.descripcion or e.cod_articulo or "Artículo").strip()
        cant = e.cantidad if e.cantidad is not None else 1
        partes.append(
            f"• Nro Pedido:  {ped} • Artículos: {cant} - {desc}"
        )
    return "_x000D_\n".join(partes) if partes else "• Nro Pedido:  0 • Artículos: 1 - Sin detalle"


def _direccion_desde_envio(env: Envio) -> str:
    loc = (env.localidad or "").strip()
    prov = (env.provincia or "BUENOS AIRES").strip()
    return f"{loc} - {prov}" if loc else prov


def _casos_abril(db: Session) -> dict[str, list[Envio]]:
    envios = list(db.scalars(select(Envio)).all())
    grupos: dict[str, list[Envio]] = defaultdict(list)
    for e in envios:
        if not es_envio_mundo2(e):
            continue
        fe = parse_fecha_tango(e.fecha_entrega) or parse_fecha_tango(e.fecha_pedido)
        if not fe or fe.year != 2026 or fe.month != 4:
            continue
        if not e.nro_pedido:
            continue
        k = clave_agrupacion_caso(e)
        if k:
            grupos[k].append(e)
    return grupos


def _muestra_estratificada(
    grupos: dict[str, list[Envio]],
    limite: int,
    seed: int,
) -> list[tuple[str, list[Envio]]]:
    rng = random.Random(seed)
    por_dia: dict[date, list[tuple[str, list[Envio]]]] = defaultdict(list)
    for k, lineas in grupos.items():
        e0 = lineas[0]
        fe = parse_fecha_tango(e0.fecha_entrega) or parse_fecha_tango(e0.fecha_pedido)
        if fe:
            por_dia[fe].append((k, lineas))

    dias = sorted(por_dia.keys())
    if not dias:
        return []

    elegidos: list[tuple[str, list[Envio]]] = []
    idx = 0
    while len(elegidos) < limite and dias:
        d = dias[idx % len(dias)]
        candidatos = por_dia[d]
        if candidatos:
            elegidos.append(candidatos.pop(rng.randrange(len(candidatos))))
            if not candidatos:
                por_dia.pop(d, None)
                dias = sorted(por_dia.keys())
        idx += 1
        if idx > limite * 30:
            break
    return elegidos[:limite]


def construir_filas(
    muestra: list[tuple[str, list[Envio]]],
    *,
    seed: int = 202604,
) -> list[dict]:
    rng = random.Random(seed)
    filas: list[dict] = []
    base_id = 900_001

    for i, (_k, lineas) in enumerate(muestra):
        e = lineas[0]
        fe_ent = parse_fecha_tango(e.fecha_entrega) or parse_fecha_tango(e.fecha_pedido)
        fe_sol = fe_ent - timedelta(days=rng.randint(1, 5)) if fe_ent else None
        local = (e.sucursal_cc or e.origen_cd or rng.choice(LOCALES) or "SF").strip()[:2].upper()
        if len(local) != 2:
            local = rng.choice(LOCALES)

        importe_cli = rng.choice([30000, 37500, 45000, 60000, 0])
        filas.append(
            {
                "idFlete": base_id + i,
                "Cliente": _cliente_desde_envio(e),
                "SolicitadoPor": "Local",
                "FechaSolicitado": _fmt_fecha(fe_sol),
                "LocalCompra": local,
                "FechaEntrega": _fmt_fecha(fe_ent),
                "LocalEntrega": local,
                "Abona": "Cliente",
                "Motivo": "ABONA CLIENTE",
                "Proveedor": "Local",
                "ImporteWamaro": 0,
                "ImporteCliente": importe_cli,
                "EstadoFlete": "Aprobado",
                "Direccion": _direccion_desde_envio(e),
                "Comentario": "SIMULACIÓN ABR 2026 — generado desde Tango importado",
                "Fletero": _elegir_fletero(rng),
                "Articulos": _articulos_desde_grupo(lineas),
            }
        )
    return filas


def escribir_excel(filas: list[dict], salida: Path, plantilla: Path | None) -> None:
    salida.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(filas)
    cols = [
        "idFlete",
        "Cliente",
        "SolicitadoPor",
        "FechaSolicitado",
        "LocalCompra",
        "FechaEntrega",
        "LocalEntrega",
        "Abona",
        "Motivo",
        "Proveedor",
        "ImporteWamaro",
        "ImporteCliente",
        "EstadoFlete",
        "Direccion",
        "Comentario",
        "Fletero",
        "Articulos",
    ]
    df = df[cols]

    if plantilla and plantilla.is_file():
        wb = load_workbook(plantilla)
        if "Resultados" in wb.sheetnames:
            del wb["Resultados"]
        ws = wb.create_sheet("Resultados", 1)
        ws.append(cols)
        for row in df.itertuples(index=False):
            ws.append(list(row))
        wb.save(salida)
    else:
        with pd.ExcelWriter(salida, engine="openpyxl") as writer:
            pd.DataFrame({"Info": ["Simulación fleteros ABR 2026"]}).to_excel(
                writer, sheet_name="Inicio", index=False
            )
            df.to_excel(writer, sheet_name="Resultados", index=False)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generar Excel simulado fleteros ABR 2026")
    parser.add_argument("--limite", type=int, default=100, help="Cantidad de solicitudes (default 100)")
    parser.add_argument("--salida", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--seed", type=int, default=202604)
    parser.add_argument("--plantilla", type=Path, default=MAYO_REF)
    args = parser.parse_args()

    if not DB_PATH.is_file():
        raise SystemExit(f"No existe la base: {DB_PATH}. Importá Tango de abril primero.")

    engine = create_engine(f"sqlite:///{DB_PATH.as_posix()}")
    with Session(engine) as db:
        grupos = _casos_abril(db)

    if not grupos:
        raise SystemExit("Sin casos Mundo 2 en abril 2026 en la base.")

    muestra = _muestra_estratificada(grupos, args.limite, args.seed)
    filas = construir_filas(muestra, seed=args.seed)
    plantilla = args.plantilla if args.plantilla.is_file() else None
    escribir_excel(filas, args.salida, plantilla)

    from collections import Counter

    dist = Counter(r["Fletero"] for r in filas)
    print(f"OK: {len(filas)} filas -> {args.salida}")
    print(f"Casos abril disponibles: {len(grupos)}")
    print("Fleteros:", dict(dist))
    print("\nProximo paso: Configuracion -> Fleteros locales -> importar este archivo -> Machear")


if __name__ == "__main__":
    main()
