"""Import y macheo colaborativo de planillas cross (Retirado por …)."""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote

import httpx
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.config import CROSS_PLANILLAS_DRIVE
from app.models import CrossSeguimiento, Envio, ImportBatch
from app.services.cross_parser import listar_hojas_workbook, parse_cross_workbook
from app.services.remito_utils import normalizar_remito


def _upsert_seguimiento(db: Session, row: dict[str, Any], batch_id: int) -> tuple[bool, bool]:
    """Returns (inserted, updated)."""
    norm = row["remito_norm"]
    existente = db.scalar(select(CrossSeguimiento).where(CrossSeguimiento.remito_norm == norm))
    if existente is None:
        db.add(
            CrossSeguimiento(
                remito_norm=norm,
                remito=row.get("remito"),
                nro_pedido=row.get("nro_pedido"),
                cod_cliente=row.get("cod_cliente"),
                importe_facturado=row.get("importe_facturado"),
                proveedor=row.get("proveedor"),
                hoja_origen=row.get("hoja_origen"),
                archivo_origen=row.get("archivo_origen"),
                import_batch_id=batch_id,
                fecha_retiro=row.get("fecha_retiro"),
                fecha_entrega_coord=row.get("fecha_entrega_coord"),
                entregado=row.get("entregado") or "pendiente",
                observacion=row.get("observacion"),
                match_estado="pendiente",
                raw_json=row.get("raw_json"),
            )
        )
        return True, False

    existente.remito = row.get("remito") or existente.remito
    existente.nro_pedido = row.get("nro_pedido") or existente.nro_pedido
    if row.get("cod_cliente"):
        existente.cod_cliente = row["cod_cliente"]
    if row.get("importe_facturado") is not None:
        existente.importe_facturado = row["importe_facturado"]
    existente.proveedor = row.get("proveedor") or existente.proveedor
    existente.hoja_origen = row.get("hoja_origen") or existente.hoja_origen
    existente.archivo_origen = row.get("archivo_origen") or existente.archivo_origen
    existente.import_batch_id = batch_id
    if row.get("fecha_retiro"):
        existente.fecha_retiro = row["fecha_retiro"]
    if row.get("fecha_entrega_coord"):
        existente.fecha_entrega_coord = row["fecha_entrega_coord"]
    if row.get("entregado"):
        existente.entregado = row["entregado"]
    if row.get("observacion"):
        existente.observacion = row["observacion"]
    if row.get("raw_json"):
        existente.raw_json = row["raw_json"]
    return False, True


def import_cross_workbook(
    db: Session,
    content: bytes,
    filename: str,
    *,
    ejecutar_macheo: bool = True,
    solo_retirado: bool = True,
) -> dict[str, Any]:
    filas, hojas = parse_cross_workbook(content, filename, solo_retirado=solo_retirado)
    batch = ImportBatch(
        filename=filename,
        source="cross_seguimiento",
        rows_in_file=len(filas),
    )
    db.add(batch)
    db.flush()

    insertados = actualizados = 0
    for row in filas:
        ins, upd = _upsert_seguimiento(db, row, batch.id)
        insertados += int(ins)
        actualizados += int(upd)

    batch.rows_inserted = insertados
    batch.rows_skipped = actualizados
    db.commit()

    macheo: dict[str, int] | None = None
    if ejecutar_macheo and (insertados or actualizados):
        macheo = ejecutar_macheo_cross(db)

    return {
        "batch_id": batch.id,
        "filename": filename,
        "hojas_procesadas": hojas,
        "hojas_disponibles": listar_hojas_workbook(content),
        "filas_agregadas": len(filas),
        "insertados": insertados,
        "actualizados": actualizados,
        "macheo": macheo,
        "message": (
            f"Cross: {insertados} nuevos, {actualizados} actualizados "
            f"({len(hojas)} pestaña(s) Retirado)."
        ),
    }


def ejecutar_macheo_cross(db: Session) -> dict[str, int]:
    envios_norm = set(
        db.scalars(
            select(Envio.remito_norm).where(
                Envio.remito_norm.isnot(None),
                Envio.remito_norm != "",
            )
        ).all()
    )
    registros = list(db.scalars(select(CrossSeguimiento)).all())
    en_maestro = sin_maestro = 0
    for reg in registros:
        if reg.remito_norm in envios_norm:
            reg.match_estado = "en_maestro"
            en_maestro += 1
        else:
            reg.match_estado = "sin_maestro"
            sin_maestro += 1
    db.commit()
    return {
        "procesados": len(registros),
        "en_maestro": en_maestro,
        "sin_maestro": sin_maestro,
    }


def resumen_cross(db: Session) -> dict[str, Any]:
    total = db.scalar(select(func.count()).select_from(CrossSeguimiento)) or 0
    en_maestro = (
        db.scalar(
            select(func.count()).select_from(CrossSeguimiento).where(
                CrossSeguimiento.match_estado == "en_maestro"
            )
        )
        or 0
    )
    entregado_si = (
        db.scalar(
            select(func.count()).select_from(CrossSeguimiento).where(
                CrossSeguimiento.entregado == "SI"
            )
        )
        or 0
    )
    entregado_no = (
        db.scalar(
            select(func.count()).select_from(CrossSeguimiento).where(
                CrossSeguimiento.entregado == "NO"
            )
        )
        or 0
    )
    return {
        "total": total,
        "en_maestro": en_maestro,
        "sin_maestro": max(0, total - en_maestro),
        "entregado_si": entregado_si,
        "entregado_no": entregado_no,
        "pendiente_entrega": max(0, total - entregado_si - entregado_no),
    }


def info_cross_remito(db: Session, remito_norm: str) -> dict[str, Any] | None:
    if not remito_norm:
        return None
    reg = db.scalar(
        select(CrossSeguimiento).where(CrossSeguimiento.remito_norm == remito_norm)
    )
    if not reg:
        return None
    return _cross_a_dict(reg)


def info_cross_caso(lineas: list[Envio], db: Session) -> dict[str, Any] | None:
    for ln in lineas:
        norm = ln.remito_norm or normalizar_remito(ln.remito)
        if norm:
            info = info_cross_remito(db, norm)
            if info:
                return info
    return None


def _cross_a_dict(reg: CrossSeguimiento) -> dict[str, Any]:
    return {
        "remito_norm": reg.remito_norm,
        "remito": reg.remito,
        "nro_pedido": reg.nro_pedido,
        "cod_cliente": reg.cod_cliente,
        "importe_facturado": reg.importe_facturado,
        "proveedor": reg.proveedor,
        "hoja_origen": reg.hoja_origen,
        "archivo_origen": reg.archivo_origen,
        "fecha_retiro": reg.fecha_retiro,
        "fecha_entrega_coord": reg.fecha_entrega_coord,
        "entregado": reg.entregado,
        "observacion": reg.observacion,
        "match_estado": reg.match_estado,
        "actualizado": reg.updated_at.isoformat() if reg.updated_at else None,
    }


_SHEET_ID_RE = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)")
_DRIVE_FILE_RE = re.compile(r"drive\.google\.com/file/d/([a-zA-Z0-9-_]+)")
_GID_RE = re.compile(r"gid=(\d+)")

# Export anónimo (sin OAuth): Google exige «Cualquiera con el enlace → Lector».
_DRIVE_UA = (
    "Mozilla/5.0 (compatible; TOP-Fletes/1.0; +https://github.com/TOP-SC/Fletes) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
_MIN_XLSX_BYTES = 3000


def parse_google_drive_url(url: str) -> dict[str, str]:
    """Extrae sheet_id/gid o file_id de un link compartido de Google."""
    raw = unquote((url or "").strip())
    if not raw:
        raise ValueError("URL vacía")
    sheet = _SHEET_ID_RE.search(raw)
    if sheet:
        gid = "0"
        gid_match = _GID_RE.search(raw)
        if gid_match:
            gid = gid_match.group(1)
        return {"tipo": "sheet", "sheet_id": sheet.group(1), "gid": gid}
    drive = _DRIVE_FILE_RE.search(raw)
    if drive:
        return {"tipo": "file", "file_id": drive.group(1)}
    raise ValueError(
        "No reconozco el link. Pegá una URL de Google Sheets "
        "(docs.google.com/spreadsheets/…) o un .xlsx en Drive."
    )


def export_url_google(meta: dict[str, str]) -> str:
    if meta["tipo"] == "sheet":
        sid = meta["sheet_id"]
        gid = str(meta.get("gid") or "0")
        if gid and gid != "0":
            return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx&gid={gid}"
        return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
    return f"https://drive.google.com/uc?export=download&id={meta['file_id']}"


def _interpretar_respuesta_drive(r: httpx.Response) -> tuple[bool, str]:
    """True si parece un xlsx válido; si no, mensaje legible."""
    ct = (r.headers.get("content-type") or "").lower()
    body = r.content or b""
    if r.status_code == 401:
        return False, (
            "HTTP 401 — el servidor no puede leer el archivo. "
            "En Drive: Compartir → «Cualquiera con el enlace» → Lector "
            "(no alcanza compartir solo con mails @empresa)."
        )
    if r.status_code == 403:
        return False, "HTTP 403 — acceso denegado. Revisá permisos del archivo."
    if r.status_code == 404:
        return False, "HTTP 404 — ID de planilla incorrecto en config.py."
    if r.status_code != 200:
        return False, f"HTTP {r.status_code}"
    if "html" in ct[:24] or body[:15].strip().lower().startswith(b"<!doctype"):
        return False, "Google devolvió HTML (login o permiso) en lugar del Excel."
    if len(body) < _MIN_XLSX_BYTES:
        return False, f"Archivo muy chico ({len(body)} bytes) — export vacío o pestaña gid incorrecta."
    if not body[:2] == b"PK":
        return False, "No es un .xlsx válido (falta firma ZIP/PK)."
    return True, "OK"


def descargar_bytes_drive(export_url: str, *, timeout: float = 180.0) -> tuple[bytes, str]:
    """GET anónimo a export de Drive/Sheets."""
    try:
        r = httpx.get(
            export_url,
            follow_redirects=True,
            timeout=timeout,
            headers={"User-Agent": _DRIVE_UA},
        )
    except httpx.TimeoutException as exc:
        raise ValueError(
            f"Timeout ({int(timeout)}s) — planilla muy pesada o red lenta. "
            "Probá de nuevo o importá el .xlsx manual."
        ) from exc
    ok, msg = _interpretar_respuesta_drive(r)
    if not ok:
        raise ValueError(msg)
    return r.content, msg


def descargar_planilla_por_id(
    file_id: str,
    *,
    gid: str = "0",
    timeout: float | None = None,
) -> tuple[bytes, str]:
    """
    Descarga .xlsx: primero con credenciales Google (OAuth / service account),
    si no hay o falla → export anónimo (requiere «Cualquiera con el enlace»).
    Devuelve (bytes, origen) origen = oauth_usuario|service_account|anonimo.
    """
    from app.config import settings
    from app.services.google_drive_auth import descargar_xlsx_autenticado, get_drive_credentials

    t = float(timeout if timeout is not None else settings.google_drive_timeout)
    creds, modo = get_drive_credentials()
    if creds is not None:
        try:
            content, modo_ok = descargar_xlsx_autenticado(file_id, timeout=t)
            return content, modo_ok
        except Exception:
            # Si auth falla, intentar anónimo (planillas aún públicas)
            pass

    meta = {"tipo": "sheet", "sheet_id": file_id, "gid": gid}
    export_url = export_url_google(meta)
    content, _msg = descargar_bytes_drive(export_url, timeout=t)
    return content, "anonimo"


def probar_planilla_drive(cfg: dict[str, str | bool]) -> dict[str, Any]:
    """Solo verifica si el export funciona (auth o anónimo), sin importar."""
    from app.config import settings

    label = str(cfg.get("label") or cfg.get("sheet_id") or "?")
    if not cfg.get("activo", True):
        return {"label": label, "ok": False, "motivo": "desactivada en config"}
    sid = cfg.get("sheet_id")
    if not sid:
        return {"label": label, "ok": False, "motivo": "sin sheet_id"}
    gid = str(cfg.get("gid") or "0")
    t = float(settings.google_drive_timeout)
    try:
        content, origen = descargar_planilla_por_id(str(sid), gid=gid, timeout=t)
        return {
            "label": label,
            "ok": True,
            "motivo": f"OK ({origen})",
            "sheet_id": str(sid),
            "gid": gid,
            "bytes": len(content),
            "http_status": 200,
            "origen": origen,
        }
    except httpx.TimeoutException:
        return {
            "label": label,
            "ok": False,
            "motivo": "Timeout — archivo grande o red lenta",
            "sheet_id": str(sid),
        }
    except Exception as exc:
        return {
            "label": label,
            "ok": False,
            "motivo": str(exc),
            "sheet_id": str(sid),
        }


def listar_estado_planillas_drive() -> list[dict[str, Any]]:
    return [probar_planilla_drive(cfg) for cfg in CROSS_PLANILLAS_DRIVE]


def descargar_planilla_drive(url: str) -> tuple[bytes, dict[str, str]]:
    """Descarga bytes de una planilla (auth Google si hay credenciales, si no anónimo)."""
    from app.config import settings

    meta = parse_google_drive_url(url)
    file_id = meta.get("sheet_id") or meta.get("file_id")
    if not file_id:
        raise ValueError("URL sin ID de archivo")
    content, origen = descargar_planilla_por_id(
        file_id,
        gid=str(meta.get("gid") or "0"),
        timeout=float(settings.google_drive_timeout),
    )
    meta["origen_descarga"] = origen
    return content, meta


def importar_cross_desde_url(
    db: Session,
    url: str,
    *,
    nombre: str | None = None,
    ejecutar_macheo: bool = True,
) -> dict[str, Any]:
    """Descarga un Excel desde link de Drive/Sheets e importa pestañas Retirado."""
    content, meta = descargar_planilla_drive(url)
    if nombre and nombre.strip():
        fname = nombre.strip()
    elif meta["tipo"] == "sheet":
        fname = f"cross_{meta['sheet_id'][:10]}.xlsx"
    else:
        fname = f"cross_{meta['file_id'][:10]}.xlsx"
    if not fname.lower().endswith(".xlsx"):
        fname += ".xlsx"
    out = import_cross_workbook(
        db,
        content,
        fname,
        ejecutar_macheo=ejecutar_macheo,
    )
    out["url_origen"] = url.strip()
    out["tipo_drive"] = meta["tipo"]
    out["origen_descarga"] = meta.get("origen_descarga")
    return out


def _path_ultimo_resultado() -> Path:
    from app.config import CROSS_INBOX_DIR

    CROSS_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    return CROSS_INBOX_DIR / "ultimo_resultado.json"


def guardar_ultimo_resultado_cross(payload: dict[str, Any]) -> None:
    """Persiste el último actualizar/import para mostrarlo en la UI (verde/rojo)."""
    data = dict(payload)
    data["cuando"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    path = _path_ultimo_resultado()
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def leer_ultimo_resultado_cross() -> dict[str, Any] | None:
    path = _path_ultimo_resultado()
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def estado_planillas_cross() -> dict[str, Any]:
    """
    Estado visible de las 5 planillas (siempre): ok / error / pendiente.
    Combina config + último resultado guardado.
    """
    ultimo = leer_ultimo_resultado_cross() or {}
    descargas = {str(d.get("label")): d for d in (ultimo.get("descargas") or [])}
    # Import por nombre de archivo
    imp = ultimo.get("importacion") if "importacion" in ultimo else ultimo
    resultados_imp = {
        str(r.get("nombre")): r for r in ((imp or {}).get("resultados") or [])
    }
    macheo = None
    if isinstance(imp, dict) and imp.get("macheo"):
        macheo = imp["macheo"]
    elif ultimo.get("macheo"):
        macheo = ultimo["macheo"]

    planillas: list[dict[str, Any]] = []
    for cfg in CROSS_PLANILLAS_DRIVE:
        label = str(cfg.get("label") or "?")
        fname = str(cfg.get("filename") or f"cross_{label}.xlsx")
        dl = descargas.get(label)
        if dl is None:
            estado = "pendiente"
            detalle = "Todavía no se actualizó desde Drive"
        elif dl.get("ok"):
            estado = "ok"
            detalle = (
                f"Descargado ({dl.get('bytes', 0):,} bytes"
                + (f" · {dl.get('origen')}" if dl.get("origen") else "")
                + ")"
            )
            imp_item = resultados_imp.get(fname)
            if imp_item is not None:
                if imp_item.get("ok"):
                    detalle += (
                        f" · importado {imp_item.get('insertados', 0)} nuevos / "
                        f"{imp_item.get('actualizados', 0)} actualizados"
                    )
                else:
                    estado = "error"
                    detalle = f"Descargó OK pero falló import: {imp_item.get('motivo')}"
        else:
            estado = "error"
            detalle = str(dl.get("motivo") or "Error al descargar")
        planillas.append(
            {
                "label": label,
                "archivo": fname,
                "estado": estado,
                "detalle": detalle,
                "activo": bool(cfg.get("activo", True)),
            }
        )

    return {
        "planillas": planillas,
        "cuando": ultimo.get("cuando"),
        "accion": ultimo.get("accion"),
        "message": ultimo.get("message"),
        "macheo": macheo,
        "hay_resultado": bool(ultimo),
    }


def actualizar_planillas_drive_a_inbox(
    *,
    importar: bool = True,
    matchear: bool = True,
    db: Session | None = None,
) -> dict[str, Any]:
    """
    Baja las planillas configuradas (CROSS_PLANILLAS_DRIVE) a ``data/cross_inbox``.
    Opcionalmente importa y machea (impacto en la base).

    No inventa macheo «en vivo» desde Drive: primero copia a carpeta del server.
    """
    from app.config import CROSS_INBOX_DIR, settings

    CROSS_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    t = float(settings.google_drive_timeout)
    descargas: list[dict[str, Any]] = []
    ok_dl = 0

    for cfg in CROSS_PLANILLAS_DRIVE:
        label = str(cfg.get("label") or cfg.get("sheet_id") or "?")
        if not cfg.get("activo", True):
            descargas.append({"label": label, "ok": False, "motivo": "desactivada"})
            continue
        sid = cfg.get("sheet_id")
        if not sid:
            descargas.append({"label": label, "ok": False, "motivo": "sin sheet_id"})
            continue
        gid = str(cfg.get("gid") or "0")
        fname = str(cfg.get("filename") or f"cross_{label}.xlsx")
        if not fname.lower().endswith(".xlsx"):
            fname += ".xlsx"
        try:
            content, origen = descargar_planilla_por_id(str(sid), gid=gid, timeout=t)
            dest = CROSS_INBOX_DIR / fname
            dest.write_bytes(content)
            ok_dl += 1
            descargas.append(
                {
                    "label": label,
                    "ok": True,
                    "archivo": fname,
                    "bytes": len(content),
                    "origen": origen,
                }
            )
        except Exception as exc:
            descargas.append(
                {
                    "label": label,
                    "ok": False,
                    "archivo": fname,
                    "motivo": str(exc),
                }
            )

    import_out: dict[str, Any] | None = None
    if importar and ok_dl > 0 and db is not None:
        import_out = importar_carpeta_cross(
            db,
            ejecutar_macheo=matchear,
            mover_procesados=True,
        )

    out = {
        "carpeta": str(CROSS_INBOX_DIR.resolve()),
        "descargas": descargas,
        "descargados_ok": ok_dl,
        "descargados_total": len(descargas),
        "importacion": import_out,
        "message": (
            f"Actualizar Drive→inbox: {ok_dl}/{len(descargas)} OK"
            + (
                f" · {import_out.get('message')}"
                if import_out and import_out.get("message")
                else ""
            )
        ),
    }
    guardar_ultimo_resultado_cross(
        {
            "accion": "Actualizar Drive → importar y machear"
            if importar
            else "Solo bajar a carpeta",
            **out,
        }
    )
    return out


def intentar_sync_drive(
    db: Session,
    *,
    ejecutar_macheo: bool = True,
) -> dict[str, Any]:
    """Compat: actualiza inbox + importa + machea."""
    return actualizar_planillas_drive_a_inbox(
        importar=True, matchear=ejecutar_macheo, db=db
    )


def listar_registros_cross(
    db: Session,
    *,
    limit: int = 200,
    solo_maestro: bool = False,
) -> list[dict[str, Any]]:
    q = select(CrossSeguimiento).order_by(CrossSeguimiento.updated_at.desc()).limit(limit)
    if solo_maestro:
        q = q.where(CrossSeguimiento.match_estado == "en_maestro")
    return [_cross_a_dict(r) for r in db.scalars(q).all()]


def listar_inbox_cross() -> dict[str, Any]:
    """Archivos .xlsx depositados en ``data/cross_inbox`` del servidor."""
    from app.config import CROSS_INBOX_DIR

    CROSS_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    archivos: list[dict[str, Any]] = []
    for p in sorted(CROSS_INBOX_DIR.glob("*.xlsx")):
        try:
            st = p.stat()
            archivos.append(
                {
                    "nombre": p.name,
                    "bytes": st.st_size,
                    "modificado": st.st_mtime,
                }
            )
        except OSError:
            continue
    return {
        "carpeta": str(CROSS_INBOX_DIR.resolve()),
        "archivos": archivos,
        "cantidad": len(archivos),
    }


def importar_carpeta_cross(
    db: Session,
    *,
    ejecutar_macheo: bool = True,
    mover_procesados: bool = True,
) -> dict[str, Any]:
    """
    Importa todos los .xlsx de ``data/cross_inbox``.
    Opcionalmente mueve cada archivo a ``data/cross_inbox/procesados/``.
    """
    from datetime import datetime

    from app.config import CROSS_INBOX_DIR

    CROSS_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    dest_dir = CROSS_INBOX_DIR / "procesados"
    if mover_procesados:
        dest_dir.mkdir(parents=True, exist_ok=True)

    resultados: list[dict[str, Any]] = []
    total_ins = total_upd = 0
    archivos = sorted(CROSS_INBOX_DIR.glob("*.xlsx"))
    for path in archivos:
        try:
            content = path.read_bytes()
            out = import_cross_workbook(
                db,
                content,
                path.name,
                ejecutar_macheo=False,
            )
            total_ins += int(out.get("insertados") or 0)
            total_upd += int(out.get("actualizados") or 0)
            item = {"nombre": path.name, "ok": True, **out}
            if mover_procesados:
                stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                dest = dest_dir / f"{stamp}_{path.name}"
                path.replace(dest)
                item["movido_a"] = dest.name
            resultados.append(item)
        except Exception as exc:
            resultados.append({"nombre": path.name, "ok": False, "motivo": str(exc)})

    macheo = ejecutar_macheo_cross(db) if ejecutar_macheo and resultados else None
    ok_n = sum(1 for x in resultados if x.get("ok"))
    out = {
        "carpeta": str(CROSS_INBOX_DIR.resolve()),
        "resultados": resultados,
        "insertados": total_ins,
        "actualizados": total_upd,
        "macheo": macheo,
        "message": (
            f"Carpeta inbox: {ok_n}/{len(resultados)} archivos OK · "
            f"{total_ins} nuevos · {total_upd} actualizados"
        ),
    }
    # Conservar descargas previas si existen, para no perder verde/rojo de Drive
    prev = leer_ultimo_resultado_cross() or {}
    guardar_ultimo_resultado_cross(
        {
            "accion": "Importar carpeta + machear",
            "descargas": prev.get("descargas") or [],
            "importacion": out,
            "macheo": macheo,
            "message": out["message"],
        }
    )
    return out


def export_cross_control_xlsx(
    db: Session,
    *,
    proveedor: str | None = None,
) -> bytes:
    """
    Planilla de control Alfaro/Fransof: cross + datos maestro
    (costo control, facturado, dif, suc, COD CLIENTE).
    """
    from io import BytesIO

    import pandas as pd
    from openpyxl.styles import Font, PatternFill

    from app.services.money_utils import EXCEL_NUM_FMT_PESOS, aplicar_formato_moneda_hoja
    from app.services.rules_service import resolver_sucursal_cc

    q = select(CrossSeguimiento).order_by(CrossSeguimiento.remito)
    if proveedor:
        q = q.where(CrossSeguimiento.proveedor == proveedor.upper())
    regs = list(db.scalars(q).all())

    # Índice maestro por remito_norm (una fila representativa)
    envios = list(
        db.scalars(
            select(Envio).where(
                Envio.remito_norm.isnot(None),
                Envio.remito_norm != "",
            )
        ).all()
    )
    by_norm: dict[str, list[Envio]] = {}
    for e in envios:
        by_norm.setdefault(e.remito_norm or "", []).append(e)

    filas: list[dict[str, Any]] = []
    for reg in regs:
        grupo = by_norm.get(reg.remito_norm or "", [])
        base = grupo[0] if grupo else None
        costo_control = None
        if grupo:
            # max por remito (mismo criterio que exports provincia)
            costo_control = max(float(e.costo_tarifario or 0) for e in grupo)
        prec_neto = None
        if grupo:
            for e in grupo:
                if e.prefactura_proveedor is not None:
                    prec_neto = float(e.prefactura_proveedor)
                    break
        facturado = reg.importe_facturado
        if facturado is None and prec_neto is not None:
            facturado = prec_neto
        control = round(costo_control, 2) if costo_control else None
        dif = None
        if facturado is not None and control is not None:
            dif = round(float(facturado) - float(control), 2)

        suc = ""
        cod_cli = reg.cod_cliente or ""
        if base:
            suc = resolver_sucursal_cc(base) or base.sucursal_cc or ""
            if not cod_cli:
                cod_cli = base.cod_cliente or ""

        filas.append(
            {
                "remito": reg.remito,
                "proveedor": reg.proveedor,
                "entregado": reg.entregado,
                "fecha_retiro": reg.fecha_retiro,
                "fecha_entrega_coord": reg.fecha_entrega_coord,
                "match_estado": reg.match_estado,
                "nro_pedido": reg.nro_pedido or (base.nro_pedido if base else None),
                "COD CLIENTE": cod_cli,
                "suc": suc,
                "destinatario": base.razon_social if base else None,
                "localidad": base.localidad if base else None,
                "provincia": base.provincia if base else None,
                "facturado": facturado,
                "control": control,
                "dif": dif,
                "observacion": reg.observacion,
                "archivo_origen": reg.archivo_origen,
                "hoja_origen": reg.hoja_origen,
            }
        )

    buf = BytesIO()
    df = pd.DataFrame(filas)
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Control cross")
        ws = writer.sheets["Control cross"]
        fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        font = Font(bold=True, color="1F4E79")
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(1, col)
            cell.fill = fill
            cell.font = font
        aplicar_formato_moneda_hoja(ws, list(df.columns))
        # Colorear entregado pendiente/NO
        for r_idx, row in enumerate(filas, start=2):
            ent = (row.get("entregado") or "").upper()
            cell = ws.cell(r_idx, list(df.columns).index("entregado") + 1)
            if ent == "NO":
                cell.fill = PatternFill(
                    start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
                )
            elif ent == "PENDIENTE":
                cell.fill = PatternFill(
                    start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"
                )
    buf.seek(0)
    return buf.getvalue()
