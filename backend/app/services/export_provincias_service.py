"""Export ARCA / costos de flete — formato contable Marcela (Excel Adrián 2025).

Replica la estructura de ``fletes año 2025.xlsx``:
  - Hoja ``Datos de la Empresa`` (WAMARO)
  - Hoja ``Listado por Imputación Contable`` con mismas columnas,
    columnas ocultas, ``-`` en provincias vacías y color por proveedor.

Los montos salen de la app (envíos / costo tarifario), no del Excel de ejemplo.
"""

from __future__ import annotations

import unicodedata
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import Envio
from app.proveedores import normalizar_proveedor
from app.services.money_utils import EXCEL_NUM_FMT_PESOS
from app.services.rules_service import es_amba_gba

# Columnas del listado Adrián (orden exacto)
_HEADERS = [
    "Cliente",  # A hidden
    "Cuenta",
    "Fecha",
    "Modelo",
    "Tipo comprobante",
    "Descripcion",  # F hidden
    "Numero Comprobante",
    "Proveedor",
    "Razón social",  # I hidden
    "Exportado",
    "DescAsientoModelo",
    "Débitos",
    "Debe",
    "Créditos",
    "Haber",
    "Saldo",
    "ImporteAlterDebe",
    "ImporteAlterHaber",
    "ImporteAlterTotal",
    "codCuenta",
    "Saldo acumulado",
    "RunningTotalAlter",
    "Barra",
    "CodModelo",
    "DescModelo",
    "Anulación",
    "DescripcionCuenta",
    "CodTurno",
    "DescTurno",
    "Puesto",
    "DescPuesto",
    "TransferidoACn",
    "Concepto",
    "CABA",
    "Buenos Aires",
    "Catamarca",
    "Chaco",
    "Chubut",
    "Córdoba",
    "Corrientes",
    "Entre Ríos",
    "Formosa",
    "Jujuy",
    "La Pampa",
    "La Rioja",
    "Mendoza",
    "Misiones",
    "Neuquén",
    "Río Negro",
    "Salta",
    "San Juan",
    "San Luis",
    "Santa Cruz",
    "Santa Fe",
    "Santiago del Estero",
    "Tierra del Fuego",
    "Tucumán",
]

_HIDDEN_COLS = {
    "A",
    "F",
    "I",
    "J",
    "K",
    "L",
    "N",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
}

_PROVINCIAS = _HEADERS[33:]  # desde CABA

# Código corto / razón social / color (como Excel Adrián)
_PROVEEDOR_META: dict[str, tuple[str, str, str | None]] = {
    "CLICPAQ": ("CLIPAQ", "Clicpaq Sa", "FFCC99FF"),
    "FRANSOF": ("FRANSO", "FRANSOF SRL", "FFCCFF66"),
    "ALFARO": ("AALFAR", "ALFARO CRISTIAN FRANCISCO ORLANDO", "FF6699FF"),
    "LBO": ("BEARIN", "LOAD BEARING OPS S.R.L.", "FFFFFF99"),
    "FLETES_SUC": ("FLETES", "Fletes sucursales AMBA/GBA", "FFFFC000"),
    "ORO NEGRO": ("ORONEG", "Oro Negro", "FF00B050"),
    "LA COSTA": ("COSTA", "Expreso La Costa", "FF00B0F0"),
}

_CUENTA = "520145 - Fletes sucursales"
_COD_CUENTA = "520145"
_MODELO = "10"
_DESC_MODELO = "10- FLETES"

_FILL_HEADER = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_FONT_HEADER = Font(bold=True, color="1F4E79")


def _norm_txt(value: str | None) -> str:
    v = (value or "").strip().upper()
    return "".join(
        c for c in unicodedata.normalize("NFD", v) if unicodedata.category(c) != "Mn"
    )


def _provincia_col(
    provincia: str | None,
    localidad: str | None,
    cp: str | None,
    *,
    excluir_planilla: bool,
) -> str:
    """Mapea destino app → columna del Excel Adrián (CABA / Buenos Aires / provincias)."""
    if excluir_planilla or es_amba_gba(provincia, localidad, cp):
        prov_u = _norm_txt(provincia)
        loc_u = _norm_txt(localidad)
        if "CABA" in prov_u or "CAPITAL FEDERAL" in prov_u or "CABA" in loc_u:
            return "CABA"
        return "Buenos Aires"

    n = _norm_txt(provincia)
    mapa = {
        "BUENOS AIRES": "Buenos Aires",
        "GRAN BUENOS AIRES": "Buenos Aires",
        "CABA": "CABA",
        "CATAMARCA": "Catamarca",
        "CHACO": "Chaco",
        "CHUBUT": "Chubut",
        "CORDOBA": "Córdoba",
        "CORRIENTES": "Corrientes",
        "ENTRE RIOS": "Entre Ríos",
        "FORMOSA": "Formosa",
        "JUJUY": "Jujuy",
        "LA PAMPA": "La Pampa",
        "LA RIOJA": "La Rioja",
        "MENDOZA": "Mendoza",
        "MISIONES": "Misiones",
        "NEUQUEN": "Neuquén",
        "RIO NEGRO": "Río Negro",
        "SALTA": "Salta",
        "SAN JUAN": "San Juan",
        "SAN LUIS": "San Luis",
        "SANTA CRUZ": "Santa Cruz",
        "SANTA FE": "Santa Fe",
        "SANTIAGO DEL ESTERO": "Santiago del Estero",
        "TIERRA DEL FUEGO": "Tierra del Fuego",
        "TUCUMAN": "Tucumán",
    }
    for key, label in mapa.items():
        if key in n:
            return label
    return "Buenos Aires"


def _meta_proveedor(raw: str | None) -> tuple[str, str, str | None]:
    key = normalizar_proveedor(raw) or (raw or "").strip().upper() or "SIN_ASIGNAR"
    if key in _PROVEEDOR_META:
        return _PROVEEDOR_META[key]
    # Fleteros locales / otros: código acotado + sin color fijo
    code = (key[:6] if key else "SIN").replace(" ", "")
    return (code, raw or key or "Sin proveedor", None)


def _filas_desde_db(db: Session) -> list[dict[str, Any]]:
    """
    Agrega costo por proveedor + mes (análogo a facturas del listado Adrián)
    y reparte montos en columnas de provincia.
    """
    q = (
        select(
            Envio.remito_norm,
            Envio.provincia,
            Envio.localidad,
            Envio.cp,
            Envio.excluir_planilla,
            Envio.proveedor_tarifa,
            Envio.fecha_entrega_d,
            func.max(func.coalesce(Envio.costo_tarifario, 0.0)).label("costo"),
        )
        .where(Envio.remito_norm.isnot(None), Envio.remito_norm != "")
        .group_by(
            Envio.remito_norm,
            Envio.provincia,
            Envio.localidad,
            Envio.cp,
            Envio.excluir_planilla,
            Envio.proveedor_tarifa,
            Envio.fecha_entrega_d,
        )
    )
    rows = db.execute(q).all()

    # key = (cod_proveedor, año-mes) -> {prov: monto, fecha_max, razon, color}
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for rem, prov, loc, cp, excl, prov_tarifa, f_ent, costo in rows:
        monto = float(costo or 0)
        if monto <= 0:
            continue
        code, razon, color = _meta_proveedor(prov_tarifa)
        periodo = ""
        fecha_ref: date | None = f_ent
        if f_ent:
            periodo = f"{f_ent.year:04d}{f_ent.month:02d}"
        else:
            periodo = "000000"
        key = (code, periodo)
        b = buckets.get(key)
        if b is None:
            b = {
                "code": code,
                "razon": razon,
                "color": color,
                "periodo": periodo,
                "fecha": fecha_ref,
                "por_prov": defaultdict(float),
                "total": 0.0,
            }
            buckets[key] = b
        col = _provincia_col(prov, loc, cp, excluir_planilla=bool(excl))
        b["por_prov"][col] += monto
        b["total"] += monto
        if fecha_ref and (b["fecha"] is None or fecha_ref > b["fecha"]):
            b["fecha"] = fecha_ref

    out: list[dict[str, Any]] = []
    for b in sorted(
        buckets.values(),
        key=lambda x: (x["fecha"] or date.min, x["code"]),
    ):
        fecha = b["fecha"] or date.today()
        nro = f"CTRL-{b['code']}-{b['periodo'] or fecha.strftime('%Y%m')}"
        fila: dict[str, Any] = {
            "Cliente": None,
            "Cuenta": _CUENTA,
            "Fecha": datetime(fecha.year, fecha.month, fecha.day),
            "Modelo": _MODELO,
            "Tipo comprobante": "FAC",
            "Descripcion": "FACTURA",
            "Numero Comprobante": nro,
            "Proveedor": b["code"],
            "Razón social": b["razon"],
            "Exportado": "Si",
            "DescAsientoModelo": _DESC_MODELO,
            "Débitos": round(b["total"], 2),
            "Debe": round(b["total"], 2),
            "Créditos": 0,
            "Haber": 0,
            "Saldo": round(b["total"], 2),
            "ImporteAlterDebe": None,
            "ImporteAlterHaber": 0,
            "ImporteAlterTotal": None,
            "codCuenta": _COD_CUENTA,
            "Saldo acumulado": round(b["total"], 2),
            "RunningTotalAlter": None,
            "Barra": None,
            "CodModelo": _MODELO,
            "DescModelo": "FLETES",
            "Anulación": None,
            "DescripcionCuenta": "Fletes sucursales",
            "CodTurno": None,
            "DescTurno": None,
            "Puesto": None,
            "DescPuesto": None,
            "TransferidoACn": None,
            "Concepto": None,
            "_color": b["color"],
        }
        for p in _PROVINCIAS:
            v = b["por_prov"].get(p)
            fila[p] = round(v, 2) if v else "-"
        out.append(fila)
    return out


def _sheet_empresa(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Datos de la Empresa"
    headers = [
        "Nombre_legal",
        "Calle",
        "Numero",
        "Piso",
        "Localidad",
        "Codigo_postal",
        "C__U__I__T__",
        "Descripcion_de_actividad_D__G__I__",
        "Logotipo_de_la_empresa",
        "Departamento",
        "Provincia",
        "Id_empresa",
        "Comuna",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, value=h)
        ws.cell(1, i).font = _FONT_HEADER
        ws.cell(1, i).fill = _FILL_HEADER
    values = [
        "WAMARO S.A.",
        None,
        None,
        None,
        None,
        None,
        "30-70904090-8",
        "VENTA AL POR MENOR DE COLCHONES Y SOMIERES",
        None,
        None,
        None,
        1,
        None,
    ]
    for i, v in enumerate(values, start=1):
        ws.cell(2, i, value=v)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 48


def _sheet_listado(wb: Workbook, filas: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Listado por Imputación Contable")
    for i, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(1, i, value=h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    money_headers = {
        "Débitos",
        "Debe",
        "Créditos",
        "Haber",
        "Saldo",
        "Saldo acumulado",
        *_PROVINCIAS,
    }

    for r_idx, fila in enumerate(filas, start=2):
        color = fila.get("_color")
        fill_prov = (
            PatternFill(start_color=color, end_color=color, fill_type="solid")
            if color
            else None
        )
        for c_idx, h in enumerate(_HEADERS, start=1):
            val = fila.get(h)
            cell = ws.cell(r_idx, c_idx, value=val)
            if h == "Proveedor" and fill_prov is not None:
                cell.fill = fill_prov
            if h in money_headers and isinstance(val, (int, float)):
                cell.number_format = EXCEL_NUM_FMT_PESOS
            if h == "Fecha" and isinstance(val, datetime):
                cell.number_format = "DD/MM/YYYY"

    # Ocultar columnas como en el Excel Adrián
    for letter in _HIDDEN_COLS:
        ws.column_dimensions[letter].hidden = True

    # Anchos útiles de columnas visibles
    widths = {
        "B": 28,
        "C": 12,
        "D": 8,
        "E": 14,
        "G": 22,
        "H": 12,
        "M": 14,
        "O": 10,
        "P": 14,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    for i in range(34, 58):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{max(1, len(filas) + 1)}"
    ws.freeze_panes = "A2"


def export_costos_por_provincia(db: Session) -> bytes:
    filas = _filas_desde_db(db)
    wb = Workbook()
    _sheet_empresa(wb)
    _sheet_listado(wb, filas)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
