from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models import Envio
from app.services.maestro_service import (
    COLUMNAS_CONTROL_MAESTRO,
    MAESTRO_COLUMNAS,
    construir_maestro,
)

FILL_CONTROL = PatternFill(start_color="E2E9F4", end_color="E2E9F4", fill_type="solid")
FILL_CONTROL_CELL = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
FONT_HEADER = Font(bold=True, color="2C3E50")


def _sheet_from_filas(filas: list[dict]) -> pd.DataFrame:
    rows = [{c: f.get(c) for c in MAESTRO_COLUMNAS} for f in filas]
    return pd.DataFrame(rows, columns=MAESTRO_COLUMNAS)


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = FONT_HEADER
        if col_name in COLUMNAS_CONTROL_MAESTRO:
            cell.fill = FILL_CONTROL
            for row_idx in range(2, len(df) + 2):
                ws.cell(row=row_idx, column=col_idx).fill = FILL_CONTROL_CELL


def export_maestro_wamaro(
    envios: list[Envio],
    *,
    incluir_excluidos: bool = True,
    db: Session | None = None,
    proveedor: str | None = None,
) -> bytes:
    tarifario_ctx = None
    if db is not None:
        from app.services.fletes_km_service import preparar_contexto_km
        from app.services.tarifario_version_service import TarifarioContext

        preparar_contexto_km(db, envios, enrich_limit=0, auto_calc_limit=0)
        tarifario_ctx = TarifarioContext(db)
    filas = construir_maestro(
        envios,
        incluir_excluidos=incluir_excluidos,
        proveedor=proveedor,
        db=db,
        tarifario_ctx=tarifario_ctx,
    )
    tort = [f for f in filas if f.get("_origen_planilla") == "tortuguitas"]
    sa = [f for f in filas if f.get("_origen_planilla") == "sa"]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_sheet(writer, _sheet_from_filas(tort), "Wamaro Tortuguitas")
        _write_sheet(writer, _sheet_from_filas(sa), "Wamaro Sa")
    buf.seek(0)
    return buf.getvalue()


def export_planilla_interior(db: Session, *, incluir_excluidos: bool = False) -> bytes:
    """Compat: export maestro WAMARO."""
    from sqlalchemy import select

    envios = list(db.scalars(select(Envio)).all())
    return export_maestro_wamaro(envios, incluir_excluidos=incluir_excluidos)
